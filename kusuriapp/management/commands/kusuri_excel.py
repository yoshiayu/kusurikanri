from django.core.management.base import BaseCommand
import openpyxl
import os
from pathlib import Path
from kusuriapp.models import CompanyMedicineName


class Command(BaseCommand):

    def handle(self, *args, **kwargs):
        medicine_object_list = []
        medicine_name_list = []

        for folder, subFolder, files in os.walk('static'):

            for file in files:
                if file.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(f'{folder}/{file}')

                    for sheet_name in wb.sheetnames:  # シートでループ
                        ws = wb[sheet_name]

                        # 一行目はヘッダーなのでスキップし、行でループ
                        for row in ws.iter_rows(min_row=2):
                            # print(sheet_name)

                            medicine = CompanyMedicineName()

                            medicine.initials = str(
                                Path(f'{folder}/{file}').parent).split('/')[-1][0]
                            medicine.company_name = sheet_name
                            # print(str(Path(f'{folder}/{file}').parent).split('/')[-1][0])
                            # print(str(Path(f'{folder}/{file}').parent).split('/')[-1][0])

                            for cell in row:  # セルでループ
                                # print(f'row: {cell.row}, column: {cell.column}, value: {cell.value}')
                                if cell.column == 2:
                                    medicine.medicine_name = cell.value
                            # print(cell.value)

                            # medicine.save()
                            # 重複していないリストを作成

                            if medicine.medicine_name not in medicine_name_list:
                                medicine_object_list.append(medicine)
                                medicine_name_list.append(
                                    medicine.medicine_name)

                        # ここで保存
                        # for medicine in medicinename_list:
                        #    medicine.save()
        CompanyMedicineName.object.bulk_create(
            medicine_object_list)


#medicine = [cell.value]
#medicine_k = []
# for i in medicine:
#    if i not in medicine_k:
#        medicine_k.append(i)

#cell.value = []

# for Q in range(ws.max_row + 1):
#    if Q == 0:
#        continue
#list = ws.cell(i, 1).value
#list_Num = Q

# for i in reversed(range(ws.max_row + 1)):
#    if i == 0:
#        break
#    if ws.cell(i, 1).value == list:
#        if i == Q:
#            continue
#    else:
#        ws.delete_row(i)
# wb.save(medicine.medicine_name)
